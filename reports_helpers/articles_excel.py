"""Excel (.xlsx) export of a review session's articles.

The column catalogue below is the single source of truth for the download popup:
the review screen fetches it to build its field checkboxes, and the same order is
used for the sheet's columns, so a field added here shows up in both places.
"""
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Confidences are stored as 0–1 floats but exported as the 0–100 percents the
# review table shows.
_CONFIDENCE_FIELDS = frozenset(
    ("relevancy_confidence", "section_category_confidence", "sentiment_confidence", "theme_confidence")
)

# Excel refuses a longer cell value than this, so long bodies/reasons are cut.
_MAX_CELL = 32767

# The columns the download popup offers, in sheet order. `default` seeds the
# popup's ticked set — the body text is offered but off by default because one of
# its cells dwarfs every other column in the file.
EXPORT_FIELDS: list[dict[str, Any]] = [
    {"key": "id", "label": "ID", "default": True},
    {"key": "title", "label": "Title", "default": True},
    {"key": "is_subscription", "label": "Subscription", "default": True},
    {"key": "is_relevant", "label": "Relevant", "default": True},
    {"key": "relevancy_confidence", "label": "Relevancy Confidence", "default": True},
    {"key": "relevancy_reason", "label": "Relevancy Reason", "default": True},
    {"key": "summary", "label": "Summary", "default": True},
    {"key": "domain_name", "label": "Publication", "default": True},
    {"key": "url", "label": "URL", "default": True},
    {"key": "date", "label": "Date", "default": True},
    {"key": "author", "label": "Author", "default": True},
    {"key": "reach", "label": "Reach", "default": True},
    {"key": "keyword_matched", "label": "Keyword Matched", "default": True},
    {"key": "section", "label": "Section", "default": True},
    {"key": "section_category_confidence", "label": "Section Confidence", "default": True},
    {"key": "section_reason", "label": "Section Reason", "default": True},
    {"key": "sentiment", "label": "Sentiment", "default": True},
    {"key": "sentiment_confidence", "label": "Sentiment Confidence", "default": True},
    {"key": "xai_sentiment_reason", "label": "Sentiment Reason", "default": True},
    {"key": "theme", "label": "Theme", "default": True},
    {"key": "theme_confidence", "label": "Theme Confidence", "default": True},
    {"key": "xai_theme_reason", "label": "Theme Reason", "default": True},
    {"key": "brand_of_interest", "label": "Brand of interest", "default": True},
    {"key": "competitors", "label": "Competitors", "default": True},
    {"key": "other_competitors", "label": "Other Competitors", "default": False},
    {"key": "peoples", "label": "People", "default": True},
    {"key": "countries", "label": "Countries", "default": True},
    {"key": "organizations", "label": "Organizations", "default": True},
    {"key": "priority_watch", "label": "Priority", "default": True},
    {"key": "syndication_of", "label": "Syndication of", "default": True},
    {"key": "similar_group_id", "label": "Story group", "default": True},
    {"key": "added_type", "label": "Added Type", "default": True},
    {"key": "is_approved_for_monitoring", "label": "Approved (Monitoring)", "default": True},
    {"key": "is_approved", "label": "Approved (Dashboards)", "default": True},
    {"key": "content", "label": "Article Text", "default": False},
]

_FIELD_LABELS = {f["key"]: f["label"] for f in EXPORT_FIELDS}
DEFAULT_EXPORT_FIELD_KEYS = [f["key"] for f in EXPORT_FIELDS if f["default"]]

# Boolean-ish fields exported as Yes/No rather than TRUE/FALSE.
_BOOL_FIELDS = frozenset(
    ("is_relevant", "is_subscription", "priority_watch", "is_approved", "is_approved_for_monitoring")
)

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def resolve_field_keys(keys: list[str] | None) -> list[str]:
    """Known export keys in catalogue order (empty/unknown input falls back to the defaults).

    Args:
        keys: Requested field keys, in any order.

    Returns:
        List of field keys ordered as in EXPORT_FIELDS.
    """
    wanted = {k for k in (keys or []) if k in _FIELD_LABELS}
    if not wanted:
        return list(DEFAULT_EXPORT_FIELD_KEYS)
    return [f["key"] for f in EXPORT_FIELDS if f["key"] in wanted]


def _cell_value(article: dict[str, Any], key: str) -> Any:
    """One article field as an Excel-safe cell value.

    Args:
        article: The canonical article dict.
        key: Export field key.

    Returns:
        A number, or a string (truncated and stripped of characters Excel rejects).
    """
    value = article.get(key)
    if key == "is_relevant":
        # A row without the flag is relevant (older articles predate the gate).
        value = value is not False
    if key in _BOOL_FIELDS:
        return "Yes" if value else "No"
    if key in _CONFIDENCE_FIELDS:
        return round(value * 100) if isinstance(value, (int, float)) else ""
    if key == "reach":
        return value if isinstance(value, (int, float)) else ""
    if isinstance(value, (list, tuple)):
        value = ", ".join(str(v) for v in value if v not in (None, ""))
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = ILLEGAL_CHARACTERS_RE.sub("", str(value))
    return text[:_MAX_CELL]


def build_articles_excel(articles: list[dict[str, Any]], field_keys: list[str] | None = None) -> bytes:
    """Build a one-sheet .xlsx of the given articles.

    Args:
        articles: Canonical article dicts, in the order they should appear.
        field_keys: Export field keys to include; empty/None uses the default set.

    Returns:
        The workbook as bytes.
    """
    keys = resolve_field_keys(field_keys)

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"

    ws.append([_FIELD_LABELS[k] for k in keys])
    for article in articles:
        ws.append([_cell_value(article, k) for k in keys])

    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Width from the widest of the first 200 rows — enough to look right without
    # walking every cell of a large export.
    for idx, key in enumerate(keys, start=1):
        widest = max(
            [len(_FIELD_LABELS[key])] + [len(str(_cell_value(a, key))) for a in articles[:200]]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(widest + 2, 12), 60)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
