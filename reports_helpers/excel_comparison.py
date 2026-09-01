"""Compare a delivered daily-report .xlsx against what the tool actually tagged.

The report workbook has one sheet per report section, each with a `Headline` /
`Media Type` header row. The article URL is read from a `URL` column when the sheet
has one, falling back to a link carried on the headline cell itself — either a real
cell hyperlink (openpyxl's `cell.hyperlink`) or a `=HYPERLINK("url","text")`
formula. All three appear in the wild; a headline styled blue-and-underlined with
no link target does not, so the plain URL column is the reliable source.

Matching against the DB is looser than `merge_helper.normalize_url`: report URLs
often carry tracking/preview parameters (`?v_preview=...`) the stored URL doesn't,
so a match is tried on the full canonical URL first and then on the URL with its
query string and fragment dropped. That second key can't replace the stored
`article_id` hash — it would change the identity every dedupe in the pipeline
relies on — so it is only ever computed here, for comparison.

A matched article is compared a second time on its section: the report's
`Section Name` column against the section the tagger assigned. Comparison is
case- and punctuation-insensitive, because the same section is written
"Corporate News", "corporate news" and "Corporate  News" across reports; only a
genuinely different section counts as a mismatch, and a row missing either side
is skipped rather than reported as one.
"""
from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import load_workbook

from file_helpers.merge_helper import normalize_url

_HEADLINE_HEADERS = ("headline", "title")
_MEDIA_TYPE_HEADERS = ("media type", "media_type", "type")
_URL_HEADERS = ("url", "link", "article url", "article_url")
_SECTION_HEADERS = ("section name", "section_name", "section")

# =HYPERLINK("https://...", "text") — the URL is the first quoted argument.
_HYPERLINK_FORMULA = re.compile(r'HYPERLINK\s*\(\s*"([^"]+)"', re.IGNORECASE)


def _cell_url(cell: Any) -> str:
    """URL carried by a cell as plain text, a real hyperlink, or a HYPERLINK formula.

    Args:
        cell: An openpyxl cell.

    Returns:
        The URL, or "" when the cell carries none.
    """
    if cell.hyperlink is not None and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    if isinstance(cell.value, str):
        text = cell.value.strip()
        if "HYPERLINK" in text.upper():
            match = _HYPERLINK_FORMULA.search(text)
            if match:
                return match.group(1).strip()
        if text.lower().startswith(("http://", "https://")):
            return text
    return ""


def loose_url_key(url: Any) -> str:
    """Canonical URL with its query string and fragment stripped.

    Args:
        url: A URL.

    Returns:
        The parameter-free canonical key, or "" for unusable input.
    """
    key = normalize_url(url)
    if not key:
        return ""
    key = key.split("#", 1)[0].split("?", 1)[0]
    return key.rstrip("/")


def read_report_rows(content: bytes) -> list[dict[str, str]]:
    """Read every sheet of a report workbook into headline/media-type/url rows.

    Sheets without a recognizable `Headline` column are skipped, as are rows whose
    headline cell has no hyperlink (section separators, blank padding rows).

    Args:
        content: The .xlsx file bytes.

    Returns:
        One dict per article row, with `sheet`, `headline`, `media_type`, `url`
        and `section`.
    """
    # data_only=False keeps formulas readable, which is how HYPERLINK() rows carry
    # their URL; a cached value would just give us the display text.
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    rows: list[dict[str, str]] = []

    for sheet in workbook.worksheets:
        header = next(sheet.iter_rows(min_row=1, max_row=1), ())
        columns = {
            str(cell.value).strip().lower(): cell.column
            for cell in header
            if cell.value is not None
        }
        headline_col = next((columns[h] for h in _HEADLINE_HEADERS if h in columns), None)
        if headline_col is None:
            continue
        media_col = next((columns[h] for h in _MEDIA_TYPE_HEADERS if h in columns), None)
        url_col = next((columns[h] for h in _URL_HEADERS if h in columns), None)
        section_col = next((columns[h] for h in _SECTION_HEADERS if h in columns), None)

        for row in sheet.iter_rows(min_row=2):
            headline_cell = row[headline_col - 1]
            url = _cell_url(row[url_col - 1]) if url_col is not None else ""
            if not url:
                url = _cell_url(headline_cell)
            if not url:
                continue
            media_type = ""
            if media_col is not None and row[media_col - 1].value is not None:
                media_type = str(row[media_col - 1].value).strip()
            section = ""
            if section_col is not None and row[section_col - 1].value is not None:
                section = str(row[section_col - 1].value).strip()
            rows.append(
                {
                    "sheet": sheet.title,
                    "headline": str(headline_cell.value or "").strip(),
                    "media_type": media_type,
                    "url": url,
                    "section": section,
                }
            )
    return rows


def section_key(section: Any) -> str:
    """Case- and punctuation-free form of a section name, for comparison only.

    Args:
        section: A section name.

    Returns:
        The comparison key, or "" when there is no usable name.
    """
    if not section:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(section).lower()).strip()


def compare_rows(rows: list[dict[str, str]], db_articles: list[dict[str, Any]]) -> dict[str, Any]:
    """Split report rows into found/missing, and flag the found ones tagged elsewhere.

    Report rows are deduplicated on the loose key first, so the same article listed
    under two sections counts once.

    Args:
        rows: Rows from `read_report_rows`.
        db_articles: The session's tagged articles, each with `url`, `section` and
            `is_relevant`.

    Returns:
        `total_report_articles`, `total_articles_found_in_tool`, `tagged_irrelevant`
        (found rows the tool marked not relevant), the `missing` rows, and
        `section_mismatches` — found rows whose report section differs from the
        tagged one.
    """
    # Both keys map to the same article, so a report URL carrying tracking params
    # still resolves to the row stored under the clean URL.
    by_url: dict[str, dict[str, Any]] = {}
    for article in db_articles:
        url = article.get("url")
        if not url:
            continue
        for key in (normalize_url(url), loose_url_key(url)):
            if key:
                by_url.setdefault(key, article)

    seen: set[str] = set()
    unique: list[dict[str, str]] = []
    for row in rows:
        key = loose_url_key(row["url"]) or row["url"]
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)

    missing: list[dict[str, str]] = []
    section_mismatches: list[dict[str, str]] = []
    tagged_irrelevant = 0
    for row in unique:
        match = by_url.get(normalize_url(row["url"])) or by_url.get(loose_url_key(row["url"]))
        if match is None:
            missing.append(row)
            continue
        # The article made the delivered report, so the tool calling it irrelevant
        # is a relevancy-gate miss, not a reason to exclude it from the counts.
        if match.get("is_relevant") is False:
            tagged_irrelevant += 1
        ai_section = match.get("section") or ""
        # A row missing either section says nothing about agreement, so skip it
        # rather than report an unverifiable mismatch.
        if not section_key(row.get("section")) or not section_key(ai_section):
            continue
        if section_key(row.get("section")) != section_key(ai_section):
            section_mismatches.append(
                {
                    "headline": row.get("headline", ""),
                    "url": row["url"],
                    "ai_section": ai_section,
                    "correct_section": row.get("section", ""),
                }
            )

    return {
        "total_report_articles": len(unique),
        "total_articles_found_in_tool": len(unique) - len(missing),
        "tagged_irrelevant": tagged_irrelevant,
        "missing": missing,
        "section_mismatches": section_mismatches,
    }
